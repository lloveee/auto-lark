# 整合后的自动化脚本 - Tkinter GUI 版本（增强版 v3.3）
# 作者：Grok 4
# 日期：2025-10-23
# 说明：修复鼠标滚轮滚动问题，确保所有模式支持上下滚动。修复坐标保存、确认窗口卡死问题，以及未解析引用问题（logging 和 Image）。
#       新增功能：启动时验证所有文件和图片路径，记录无效路径到日志。
#       布局：每个模式 tab 下内容分成三列（坐标、文件、图片），支持鼠标滚轮上下滚动。
#       包含高级技巧：屏幕分辨率适配、实时鼠标位置、坐标测试按钮、动态等待、日志分级、进度保存、快捷键控制。

# ========== 公共导入（所有模式共享） ==========
import pyautogui
import time
import pandas as pd
import os
import pyperclip
from openpyxl import load_workbook, Workbook
from zipfile import BadZipFile
from datetime import datetime
import traceback
import csv
import clipboard
import json
from threading import Event, Thread
import logging
import keyboard
import tkinter as tk
from tkinter import messagebox, ttk, Entry, Button, Label, Canvas, Scrollbar
from PIL import Image

# 配置日志
logging.basicConfig(
    filename=r'E:\pythonProject\automation.log',
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    encoding='utf-8'
)

# config 文件路径
config_path = r'E:\pythonProject\config.json'
progress_path = r'E:\pythonProject\progress.json'

# 默认坐标及注释
default_configs = {
    "查履约": {
        "search_box_position_lu": {"value": [58, 270], "comment": "# 左上搜索框坐标（示例，需调整）"},
        "search_box_position_ru": {"value": [1159, 241], "comment": "# 右上搜索框坐标（示例，需调整）"},
        "search_box_position_ld": {"value": [58, 778], "comment": "# 左下搜索框坐标（示例，需调整）"},
        "search_box_position_rd": {"value": [1157, 742], "comment": "# 右下搜索框坐标（示例，需调整）"},
        "lu_region": {"value": [0, 0, 960, 540], "comment": "# 左上 (left, top, width, height)"},
        "ru_region": {"value": [960, 0, 960, 540], "comment": "# 右上"},
        "ld_region": {"value": [0, 540, 960, 540], "comment": "# 左下"},
        "rd_region": {"value": [960, 540, 960, 540], "comment": "# 右下"}
    },
    "通过样品": {
        "POS_SEARCH_BOX": {"value": [714, 406], "comment": "# 搜索框"},
        "POS_SEARCH_RESULT": {"value": [739, 459], "comment": "# 搜索结果（若需点击）"},
        "POS_EXPAND_DETAIL": {"value": [1554, 556], "comment": "# 展开详情按钮"},
        "REGION_TEXT_START": {"value": [757, 566], "comment": "# 框选文字起点"},
        "REGION_TEXT_END": {"value": [846, 566], "comment": "# 框选文字终点"},
        "POS_ORDER_BTN": {"value": [859, 690], "comment": "# 点击按钮复制订单号"},
        "POS_CHAT_BTN": {"value": [808, 672], "comment": "# 达人聊天按钮"},
        "POS_SEND_BTN": {"value": [515, 823], "comment": "# send按钮（如果需要先点 send 再点输入框）"},
        "POS_CHAT_BOX": {"value": [514, 918], "comment": "# 聊天框输入位置"},
        "POS_AFTER_SEND": {"value": [391, 20], "comment": "# 发送后点击的指定位置"},
        "POS_CLICK1": {"value": [616, 553], "comment": ""},
        "POS_CLICK2": {"value": [1553, 332], "comment": ""},
        "POS_CLOSE1": {"value": [970, 20], "comment": ""},
        "POS_CLOSE2": {"value": [732, 20], "comment": ""}
    },
    "查到货": {
        "search_box": {"value": [987, 250], "comment": "# 搜索框"},
        "pick_box": {"value": [1490, 628], "comment": "# 点击物流信息按钮"},
        "date_box": {"value": [1539, 688], "comment": "# 日期文本"},
        "close_box": {"value": [1855, 973], "comment": "# ❗物流弹窗的关闭按钮（如果是弹窗）"}
    },
    "催视频码": {
        "search_box_click_pos": {"value": [683, 387], "comment": ""},
        "back_to_search_pos": {"value": [168, 19], "comment": ""},
        "message_input_pos": {"value": [510, 918], "comment": "# 备用，未使用"}
    }
}

# 模式的文件和图片路径
mode_paths = {
    "查履约": {
        "files": [
            {"name": "excel_input", "path": r"E:\pythonProject\查履约.xlsx"},
            {"name": "excel_output", "path": r"E:\pythonProject\查履约 - 副本.xlsx"}
        ],
        "images": [
            {"name": "chaoshi", "path": r"E:\pythonProject\chaoshi.png"},
            {"name": "zhengzaizuo", "path": r"E:\pythonProject\zhengzaizuo.png"},
            {"name": "yiwancheng", "path": r"E:\pythonProject\yiwancheng.png"},
            {"name": "quxiao", "path": r"E:\pythonProject\quxiao.png"},
            {"name": "xitongquxiao", "path": r"E:\pythonProject\xitongquxiao.png"},
            {"name": "yunsongzhong", "path": r"E:\pythonProject\yunsongzhong.png"}
        ]
    },
    "通过样品": {
        "files": [
            {"name": "excel_path", "path": r"E:\pythonProject\达人任务.xlsx"},
            {"name": "output_path", "path": r"E:\pythonProject\达人任务结果.xlsx"},
            {"name": "chat_message_path", "path": r"E:\pythonProject\达人信息.txt"},
            {"name": "log_file", "path": r"E:\pythonProject\send_log.txt"}
        ],
        "images": [
            {"name": "copy_btn", "path": r"E:\pythonProject\copy_btn.png"},
            {"name": "fasongbuliao", "path": r"E:\pythonProject\fasongbuliao.png"}
        ]
    },
    "查到货": {
        "files": [
            {"name": "csv_path", "path": r"E:\pythonProject\送样未到达订单号.csv"},
            {"name": "excel_path", "path": r"E:\pythonProject\送样未到达订单号日志.xlsx"}
        ],
        "images": []
    },
    "催视频码": {
        "files": [
            {"name": "orders_file", "path": r"E:\pythonProject\darenxinxi\订单号.xlsx"},
            {"name": "message_file", "path": r"E:\pythonProject\darenxinxi\message.xlsx"},
            {"name": "log_file", "path": r"E:\pythonProject\darenxinxi\运行日志.xlsx"}
        ],
        "images": [
            {"name": "liaotiananniu", "path": r"E:\pythonProject\darenxinxi\liaotiananniu.png"},
            {"name": "fasongbuliao", "path": r"E:\pythonProject\darenxinxi\fasongbuliao.png"},
            {"name": "fasongbuliao2", "path": r"E:\pythonProject\darenxinxi\fasongbuliao2.png"},
            {"name": "sendamessage", "path": r"E:\pythonProject\darenxinxi\sendamessage.png"}
        ]
    }
}

# 验证文件和图片路径
def validate_paths():
    for mode, paths in mode_paths.items():
        for img in paths["images"]:
            if not os.path.exists(img["path"]):
                logging.error(f"图片路径无效: {img['path']}")
        for file in paths["files"]:
            if not os.path.exists(file["path"]):
                logging.error(f"文件路径无效: {file['path']}")

# 加载自定义路径和坐标（从 config.json）
if os.path.exists(config_path):
    try:
        with open(config_path, 'r') as f:
            loaded_configs = json.load(f)
        for mode, data in loaded_configs.items():
            if mode in mode_paths and 'files' in data:
                for file_dict in mode_paths[mode]['files']:
                    for loaded_file in data['files']:
                        if file_dict['name'] == loaded_file['name']:
                            file_dict['path'] = loaded_file['path']
            if mode in mode_paths and 'images' in data:
                for img_dict in mode_paths[mode]['images']:
                    for loaded_img in data['images']:
                        if img_dict['name'] == loaded_img['name']:
                            img_dict['path'] = loaded_img['path']
            if mode in default_configs:
                for name, value in data.items():
                    if name in default_configs[mode] and isinstance(value, (list, tuple)):
                        default_configs[mode][name]["value"] = value
    except Exception as e:
        logging.error(f"加载 config.json 失败: {e}")

# 保存路径和坐标到 config.json
def save_paths_and_configs():
    save_data = {}
    for mode in mode_paths:
        save_data[mode] = {
            'files': [{ 'name': f['name'], 'path': f['path'] } for f in mode_paths[mode]['files']],
            'images': [{ 'name': i['name'], 'path': i['path'] } for i in mode_paths[mode]['images']],
            **{name: info["value"] for name, info in default_configs.get(mode, {}).items()}
        }
    try:
        with open(config_path, 'w') as f:
            json.dump(save_data, f, indent=4)
    except Exception as e:
        logging.error(f"保存 config.json 失败: {e}")

# 动态坐标适配
def scale_coord(coord, base_width=1920, base_height=1080):
    screen_width, screen_height = pyautogui.size()
    if len(coord) == 2:  # 点坐标 (x, y)
        return [int(coord[0] * screen_width / base_width), int(coord[1] * screen_height / base_height)]
    elif len(coord) == 4:  # 区域坐标 (left, top, width, height)
        return [
            int(coord[0] * screen_width / base_width),
            int(coord[1] * screen_height / base_height),
            int(coord[2] * screen_width / base_width),
            int(coord[3] * screen_height / base_height)
        ]
    return coord

# 重设坐标函数
def reset_coord(mode, coord_name, label):
    coord = default_configs[mode][coord_name]["value"]
    is_region = len(coord) == 4

    capture_win = tk.Toplevel()
    capture_win.attributes('-fullscreen', True)
    capture_win.attributes('-alpha', 0.01)
    capture_win.attributes('-topmost', True)

    clicks = []
    label_text = tk.Label(capture_win, text="", bg='white', fg='black')
    label_text.place(x=10, y=10)

    def update_mouse_pos():
        try:
            x, y = pyautogui.position()
            label_text.config(text=f"当前鼠标位置: ({x}, {y})")
            capture_win.after(100, update_mouse_pos)
        except tk.TclError:
            pass

    update_mouse_pos()

    def on_click(event):
        clicks.append((event.x, event.y))
        if (not is_region and len(clicks) == 1) or (is_region and len(clicks) == 2):
            if not is_region:
                new_coord = list(clicks[0])
            else:
                x1, y1 = clicks[0]
                x2, y2 = clicks[1]
                new_coord = [min(x1, x2), min(y1, y2), abs(x2 - x1), abs(y2 - y1)]
            default_configs[mode][coord_name]["value"] = new_coord
            save_paths_and_configs()
            capture_win.destroy()  # 先销毁捕获窗口
            label.config(text=f"{coord_name} = {new_coord} {default_configs[mode][coord_name]['comment']}")
            messagebox.showinfo("成功", f"已更新 {coord_name} 为 {new_coord}")

    capture_win.bind('<Button-1>', on_click)

    msg = "点击以设置位置" if not is_region else "先点击左上角，再点击右下角以设置区域"
    tk.Label(capture_win, text=msg, bg='white', fg='black').place(x=10, y=40)

# 动态等待函数
def wait_for_element(image, timeout=10, confidence=0.8):
    start = time.time()
    while time.time() - start < timeout:
        try:
            if isinstance(image, Image.Image):
                if pyautogui.locateOnScreen(image, confidence=confidence):
                    return True
            elif isinstance(image, str) and os.path.exists(image):
                if pyautogui.locateOnScreen(image, confidence=confidence):
                    return True
        except:
            pass
        time.sleep(0.5)
    return False

# 日志函数
def log(msg, level='info'):
    try:
        if level == 'info':
            logging.info(msg)
        elif level == 'error':
            logging.error(msg)
        print(msg)
    except Exception as e:
        print(f"日志记录失败: {e}")

# ========== 模式1: “查履约” 代码（支持暂停/停止） ==========
def run_mode_cha_lvyue(pause_event, stop_event):
    configs = default_configs["查履约"]
    excel_input = mode_paths["查履约"]["files"][0]["path"]
    excel_output = mode_paths["查履约"]["files"][1]["path"]
    images = {}
    for img_info in mode_paths["查履约"]["images"]:
        try:
            if os.path.exists(img_info["path"]):
                images[img_info["name"]] = Image.open(img_info["path"])
            else:
                log(f"图片 {img_info['path']} 不存在", level='error')
                images[img_info["name"]] = None
        except Exception as e:
            log(f"加载图片 {img_info['path']} 失败: {e}", level='error')
            images[img_info["name"]] = None
    images = {k: v for k, v in images.items() if v is not None}

    def click_search_box(position):
        scaled_pos = scale_coord(position)
        pyautogui.moveTo(scaled_pos[0], scaled_pos[1], duration=0)
        pyautogui.click()
        time.sleep(0.5)

    def enter_text(text):
        try:
            pyautogui.hotkey('ctrl', 'a')
            time.sleep(0.1)
            pyautogui.press('backspace')
            time.sleep(0.1)
        except Exception as e:
            log(f"清除输入框失败: {e}", level='error')
            pyautogui.click(clicks=3, interval=0.05)
            time.sleep(0.1)
            pyautogui.press('backspace')
            time.sleep(0.1)
        pyautogui.typewrite('\b' * 20)
        time.sleep(0.2)
        pyperclip.copy(text)
        pyautogui.hotkey('ctrl', 'v')
        time.sleep(0.1)
        pyautogui.press('enter')

    def recognize_status(region=None):
        scaled_region = scale_coord(region) if region else None
        for attempt in range(10):
            pause_event.wait()
            if stop_event.is_set():
                return "停止"
            for status, img in images.items():
                try:
                    location = pyautogui.locateOnScreen(img, confidence=0.7, grayscale=True, region=scaled_region)
                    if location:
                        log(f"尝试 {attempt+1}: 识别到 {status} at {location}")
                        return status
                except:
                    pass
            time.sleep(1)
        log("经过多次尝试，仍未识别到任何状态", level='error')
        return "未识别"

    def save_result(data):
        try:
            if os.path.exists(excel_output):
                try:
                    load_workbook(excel_output)
                    with pd.ExcelWriter(excel_output, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
                        data.to_excel(writer, sheet_name="结果", index=False)
                    return
                except BadZipFile:
                    log("检测到结果文件损坏，正在重建...", level='error')
                    os.remove(excel_output)
            data.to_excel(excel_output, sheet_name="结果", index=False, engine="openpyxl")
        except Exception as e:
            log(f"保存结果失败: {e}", level='error')

    try:
        df = pd.read_excel(excel_input)
    except FileNotFoundError:
        log(f"Excel文件 {excel_input} 不存在", level='error')
        return
    except ValueError:
        log(f"Excel文件格式错误", level='error')
        return
    except Exception as e:
        log(f"读取Excel失败: {e}", level='error')
        return

    results = []
    i = 0
    if os.path.exists(progress_path):
        try:
            with open(progress_path, 'r') as f:
                i = json.load(f).get('index', 0)
        except Exception as e:
            log(f"加载进度文件失败: {e}", level='error')
    for t in range(3, 0, -1):
        log(f"{t}秒后开始...")
        time.sleep(1)
    log("开始处理！")

    while i < len(df):
        pause_event.wait()
        if stop_event.is_set():
            with open(progress_path, 'w') as f:
                json.dump({'index': i}, f)
            break
        quadrants = [
            (str(df.iloc[i, 0]) if i < len(df) else None, configs["search_box_position_lu"]["value"], configs["lu_region"]["value"], "左上", i+1),
            (str(df.iloc[i+1, 0]) if i+1 < len(df) else None, configs["search_box_position_ru"]["value"], configs["ru_region"]["value"], "右上", i+2),
            (str(df.iloc[i+2, 0]) if i+2 < len(df) else None, configs["search_box_position_ld"]["value"], configs["ld_region"]["value"], "左下", i+3),
            (str(df.iloc[i+3, 0]) if i+3 < len(df) else None, configs["search_box_position_rd"]["value"], configs["rd_region"]["value"], "右下", i+4)
        ]

        for search_text, position, _, screen, idx in quadrants:
            if search_text:
                log(f"正在输入 {screen} 第 {idx} 条: {search_text}")
                click_search_box(position)
                enter_text(search_text)

        if not wait_for_element(images.get("yiwancheng"), timeout=8):
            log("页面加载超时", level='error')

        for search_text, _, region, screen, idx in quadrants:
            pause_event.wait()
            if stop_event.is_set():
                with open(progress_path, 'w') as f:
                    json.dump({'index': i}, f)
                break
            if search_text:
                log(f"正在识别 {screen} 第 {idx} 条: {search_text}")
                status = recognize_status(region=region)
                log(f"{screen} 识别结果: {status}")
                results.append({"搜索内容": search_text, "状态": status, "屏幕": screen})

        try:
            save_result(pd.DataFrame(results))
        except Exception as e:
            log(f"保存结果失败: {e}", level='error')

        time.sleep(2)
        i += 4

    if os.path.exists(progress_path):
        os.remove(progress_path)
    log("全部完成 ✅")

# ========== 模式2: “通过样品” 代码（支持暂停/停止） ==========
def run_mode_tongguo_yangpin(pause_event, stop_event):
    configs = default_configs["通过样品"]
    excel_path = mode_paths["通过样品"]["files"][0]["path"]
    output_path = mode_paths["通过样品"]["files"][1]["path"]
    screenshot_path = mode_paths["通过样品"]["images"][0]["path"]
    fasongbuliao_path = mode_paths["通过样品"]["images"][1]["path"]
    chat_message_path = mode_paths["通过样品"]["files"][2]["path"]
    log_file = mode_paths["通过样品"]["files"][3]["path"]
    MODE_INNER = 'single'

    POS_SEARCH_BOX = scale_coord(configs["POS_SEARCH_BOX"]["value"])
    POS_SEARCH_RESULT = scale_coord(configs["POS_SEARCH_RESULT"]["value"])
    POS_EXPAND_DETAIL = scale_coord(configs["POS_EXPAND_DETAIL"]["value"])
    REGION_TEXT_START = scale_coord(configs["REGION_TEXT_START"]["value"])
    REGION_TEXT_END = scale_coord(configs["REGION_TEXT_END"]["value"])
    POS_ORDER_BTN = scale_coord(configs["POS_ORDER_BTN"]["value"])
    POS_CHAT_BTN = scale_coord(configs["POS_CHAT_BTN"]["value"])
    POS_SEND_BTN = scale_coord(configs["POS_SEND_BTN"]["value"])
    POS_CHAT_BOX = scale_coord(configs["POS_CHAT_BOX"]["value"])
    POS_AFTER_SEND = scale_coord(configs["POS_AFTER_SEND"]["value"])
    POS_CLICK1 = scale_coord(configs["POS_CLICK1"]["value"])
    POS_CLICK2 = scale_coord(configs["POS_CLICK2"]["value"])
    POS_CLOSE1 = scale_coord(configs["POS_CLOSE1"]["value"])
    POS_CLOSE2 = scale_coord(configs["POS_CLOSE2"]["value"])

    WAIT_BEFORE_START = 5
    SLEEP_AFTER_SEARCHBOX = 0.5
    SLEEP_AFTER_CLEAR = 1
    SLEEP_AFTER_INPUT = 4.5
    SLEEP_AFTER_EXPAND = 2
    SLEEP_AFTER_SELECT_TEXT = 1
    SLEEP_AFTER_COPY_TEXT = 1
    SLEEP_AFTER_ORDER_BTN = 1
    SLEEP_AFTER_CHAT_BTN = 6
    SLEEP_AFTER_SEND_BTN = 2
    SLEEP_AFTER_CHAT_BOX = 1
    SLEEP_AFTER_AFTER_SEND = 1
    SLEEP_AFTER_CLICK1 = 5
    SLEEP_AFTER_CLICK2 = 3
    SLEEP_AFTER_CLICK3 = 1.5
    SLEEP_AFTER_CLOSE1 = 2
    SLEEP_AFTER_CLOSE2 = 2
    LOCATE_CONFIDENCE = 0.7

    pyautogui.FAILSAFE = True
    pyautogui.PAUSE = 0.12

    def nowstr():
        return datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    def log_to_file(msg):
        ts = nowstr()
        line = f"[{ts}] {msg}"
        try:
            with open(log_file, "a", encoding="utf-8") as lf:
                lf.write(line + "\n")
        except:
            log(f"无法写入日志文件：{log_file}", level='error')

    CHAT_MESSAGE = ""
    if MODE_INNER == 'single':
        if os.path.exists(chat_message_path):
            with open(chat_message_path, "r", encoding="utf-8") as f:
                CHAT_MESSAGE = f.read().strip()
            log(f"已载入统一聊天信息（来自 {chat_message_path}）")
        else:
            log(f"警告：MODE='single'，但未找到文件 {chat_message_path}，聊天信息将为空", level='error')

    log(f"程序将在 {WAIT_BEFORE_START} 秒后开始自动运行，请切换到目标窗口...")
    for i in range(WAIT_BEFORE_START, 0, -1):
        log(f"{i} ...")
        time.sleep(1)

    try:
        df = pd.read_excel(excel_path)
    except FileNotFoundError:
        log(f"Excel文件 {excel_path} 不存在", level='error')
        return
    except ValueError:
        log(f"Excel文件格式错误", level='error')
        return
    except Exception as e:
        log(f"读取Excel失败: {e}", level='error')
        return

    if "列1" not in df.columns: df["列1"] = ""
    if "列2" not in df.columns: df["列2"] = ""
    if "列3" not in df.columns: df["列3"] = ""
    if "发送状态" not in df.columns: df["发送状态"] = ""
    if "发送时间" not in df.columns: df["发送时间"] = ""
    if MODE_INNER == 'per_row' and "Message" not in df.columns:
        df["Message"] = ""

    total = len(df)
    log(f"开始处理，共 {total} 条记录")
    i = 0
    if os.path.exists(progress_path):
        try:
            with open(progress_path, 'r') as f:
                i = json.load(f).get('index', 0)
        except Exception as e:
            log(f"加载进度文件失败: {e}", level='error')

    for idx in range(i, len(df)):
        pause_event.wait()
        if stop_event.is_set():
            with open(progress_path, 'w') as f:
                json.dump({'index': idx}, f)
            break
        name = str(df.iloc[idx, 0]).strip()
        try:
            log(f"开始处理第 {idx+1}/{total}：{name}")
            pyautogui.click(*POS_SEARCH_BOX)
            time.sleep(SLEEP_AFTER_SEARCHBOX)
            pyautogui.hotkey("ctrl", "a")
            pyautogui.press("backspace")
            time.sleep(SLEEP_AFTER_CLEAR)
            pyperclip.copy(name)
            pyautogui.hotkey("ctrl", "v")
            pyautogui.press("enter")
            time.sleep(SLEEP_AFTER_INPUT)
            pyautogui.click(*POS_EXPAND_DETAIL)
            time.sleep(SLEEP_AFTER_EXPAND)
            pyautogui.moveTo(*REGION_TEXT_START)
            pyautogui.mouseDown()
            pyautogui.moveTo(*REGION_TEXT_END, duration=0.5)
            pyautogui.mouseUp()
            time.sleep(SLEEP_AFTER_SELECT_TEXT)
            pyautogui.hotkey("ctrl", "c")
            time.sleep(SLEEP_AFTER_COPY_TEXT)
            text = pyperclip.paste()
            df.loc[idx, "列1"] = "" if text is None else str(text).strip()
            log(f"复制到列1：{df.loc[idx,'列1'][:80]}")
            prev_clip = pyperclip.paste()
            pyautogui.click(*POS_ORDER_BTN)
            time.sleep(SLEEP_AFTER_ORDER_BTN)
            order_number = pyperclip.paste()
            if order_number == prev_clip:
                pyautogui.hotkey("ctrl", "c")
                time.sleep(0.3)
                order_number = pyperclip.paste()
            df.loc[idx, "列2"] = "" if order_number is None else str(order_number).strip()
            log(f"复制到列2（订单号）：{df.loc[idx,'列2']}")
            pyautogui.click(*POS_CHAT_BTN)
            if not wait_for_element(fasongbuliao_path, timeout=SLEEP_AFTER_CHAT_BTN):
                log("聊天页面加载超时", level='error')
            skip_send = False
            if os.path.exists(fasongbuliao_path):
                try:
                    button_location = pyautogui.locateCenterOnScreen(fasongbuliao_path, confidence=LOCATE_CONFIDENCE)
                except:
                    button_location = None
                if button_location:
                    log(f"找到 fasongbuliao.png，跳过该达人发送操作")
                    skip_send = True
                else:
                    log(f"未找到 fasongbuliao.png，继续发送")
            else:
                log(f"⚠️ 未找到图片文件：{fasongbuliao_path}，继续发送", level='error')

            if skip_send:
                pyautogui.click(*POS_AFTER_SEND)
                time.sleep(SLEEP_AFTER_AFTER_SEND)
                df.loc[idx, "发送状态"] = "跳过（发送不了）"
                df.loc[idx, "发送时间"] = nowstr()
                continue

            pyautogui.click(*POS_SEND_BTN)
            time.sleep(SLEEP_AFTER_SEND_BTN)
            pyautogui.click(*POS_CHAT_BOX)
            time.sleep(SLEEP_AFTER_CHAT_BOX)

            if MODE_INNER == 'single':
                chat_msg = CHAT_MESSAGE
            else:
                chat_msg = str(df.iloc[idx].get("Message", "") or "")

            if chat_msg.strip():
                pyperclip.copy(chat_msg)
                pyautogui.hotkey("ctrl", "v")
                time.sleep(0.25)
                pyautogui.press("enter")
                log("聊天信息已发送")
            else:
                log("聊天信息为空，跳过发送", level='error')

            pyautogui.click(*POS_AFTER_SEND)
            time.sleep(SLEEP_AFTER_AFTER_SEND)
            pyautogui.click(*POS_CLICK1)
            time.sleep(SLEEP_AFTER_CLICK1)
            pyautogui.click(*POS_CLICK2)
            time.sleep(0.5)
            time.sleep(SLEEP_AFTER_CLICK2)
            link = ""
            if os.path.exists(screenshot_path):
                try:
                    button_location = pyautogui.locateCenterOnScreen(screenshot_path, confidence=LOCATE_CONFIDENCE)
                except:
                    button_location = None
                if button_location:
                    pyautogui.click(button_location)
                    time.sleep(SLEEP_AFTER_CLICK3)
                    link = pyperclip.paste() or ""
                    df.loc[idx, "列3"] = str(link).strip()
                    log(f"识别并复制链接到列3：{df.loc[idx,'列3']}")
                else:
                    log(f"⚠️ 未找到截图按钮（{screenshot_path}），跳过复制链接", level='error')
                    df.loc[idx, "列3"] = ""
            else:
                log(f"⚠️ 未找到截图文件：{screenshot_path}，跳过 locate", level='error')

            pyautogui.click(*POS_CLOSE1)
            time.sleep(SLEEP_AFTER_CLOSE1)
            pyautogui.click(*POS_CLOSE2)
            time.sleep(SLEEP_AFTER_CLOSE2)
            df.loc[idx, "发送状态"] = "成功"
            df.loc[idx, "发送时间"] = nowstr()
            log(f"完成：{name} ✅")

        except Exception as e:
            df.loc[idx, "发送状态"] = "失败"
            df.loc[idx, "发送时间"] = nowstr()
            log(f"处理 {name} 时发生异常：{e}", level='error')
            log(traceback.format_exc(), level='error')

        finally:
            try:
                df.to_excel(output_path, index=False)
                log(f"进度已保存到 {output_path} （{idx+1}/{total}）")
            except Exception as save_err:
                log(f"保存进度失败：{save_err}", level='error')

    try:
        df.to_excel(output_path, index=False)
        log(f"全部处理完成，最终结果已保存到：{output_path}")
    except Exception as e:
        log(f"最后保存失败：{e}", level='error')

    if os.path.exists(progress_path):
        os.remove(progress_path)
    log("脚本结束。")

# ========== 模式3: “查到货” 代码（支持暂停/停止） ==========
def run_mode_cha_daohuo(pause_event, stop_event):
    configs = default_configs["查到货"]
    csv_path = mode_paths["查到货"]["files"][0]["path"]
    excel_path = mode_paths["查到货"]["files"][1]["path"]
    try:
        with open(csv_path, 'r', encoding='utf-8') as f:
            reader = csv.reader(f)
            order_numbers = [row[0].strip() for row in reader if row and row[0].strip()]
    except FileNotFoundError:
        log(f"CSV文件 {csv_path} 不存在", level='error')
        return
    except Exception as e:
        log(f"读取CSV失败: {e}", level='error')
        return

    log("读取到订单号：" + str(order_numbers))

    log("5 秒后开始，请手动打开订单搜索页面")
    time.sleep(5)

    search_box = scale_coord(configs["search_box"]["value"])
    pick_box = scale_coord(configs["pick_box"]["value"])
    date_box = scale_coord(configs["date_box"]["value"])
    close_box = scale_coord(configs["close_box"]["value"])

    wb = Workbook()
    ws = wb.active
    ws.title = "物流日期"
    ws.append(["订单号", "物流日期"])

    i = 0
    if os.path.exists(progress_path):
        try:
            with open(progress_path, 'r') as f:
                i = json.load(f).get('index', 0)
        except Exception as e:
            log(f"加载进度文件失败: {e}", level='error')

    for idx, order in enumerate(order_numbers[i:], start=i):
        pause_event.wait()
        if stop_event.is_set():
            with open(progress_path, 'w') as f:
                json.dump({'index': idx}, f)
            break
        try:
            pyautogui.click(*search_box)
            pyautogui.hotkey('ctrl', 'a')
            pyautogui.press('backspace')
            pyautogui.typewrite(order, interval=0.05)
            pyautogui.press('enter')
            time.sleep(2.5)
            pyautogui.moveTo(*pick_box)
            time.sleep(0.3)
            pyautogui.click(*pick_box)
            time.sleep(3.8)
            pyautogui.moveTo(*date_box)
            pyautogui.mouseDown()
            time.sleep(0.3)
            pyautogui.move(0, 100)
            pyautogui.mouseUp()
            pyautogui.hotkey('ctrl', 'c')
            time.sleep(0.5)
            date_str = pyperclip.paste().strip()
            try:
                date_obj = datetime.strptime(date_str, "%m/%d/%Y %I:%M:%S %p")
            except ValueError:
                try:
                    date_obj = datetime.strptime(date_str, "%Y-%m-%d %H:%M:%S")
                except:
                    date_obj = date_str
            ws.append([order, date_obj])
            log(f"✅ 订单 {order} 日期写入完成")
            wb.save(excel_path)
            if close_box:
                pyautogui.click(*close_box)
            time.sleep(2.0)
        except Exception as e:
            log(f"处理订单 {order} 失败: {e}", level='error')
            log(traceback.format_exc(), level='error')

    if os.path.exists(progress_path):
        os.remove(progress_path)
    log(f"所有订单日期已保存到 Excel：{excel_path}")

# ========== 模式4: “催视频码” 代码（支持暂停/停止） ==========
def run_mode_cui_shipinma(pause_event, stop_event):
    configs = default_configs["催视频码"]
    orders_file = mode_paths["催视频码"]["files"][0]["path"]
    message_file = mode_paths["催视频码"]["files"][1]["path"]
    log_file = mode_paths["催视频码"]["files"][2]["path"]
    image_paths = {
        "chat_button_img": mode_paths["催视频码"]["images"][0]["path"],
        "no_send_img": mode_paths["催视频码"]["images"][1]["path"],
        "no_send_img2": mode_paths["催视频码"]["images"][2]["path"],
        "send_message_img": mode_paths["催视频码"]["images"][3]["path"]
    }
    images = {}
    for name, path in image_paths.items():
        try:
            if os.path.exists(path):
                images[name] = Image.open(path)
            else:
                log(f"图片 {path} 不存在", level='error')
                images[name] = None
        except Exception as e:
            log(f"加载图片 {path} 失败: {e}", level='error')
            images[name] = None
    chat_button_img = images["chat_button_img"]
    no_send_img = images["no_send_img"]
    no_send_img2 = images["no_send_img2"]
    send_message_img = images["send_message_img"]

    search_box_click_pos = scale_coord(configs["search_box_click_pos"]["value"])
    back_to_search_pos = scale_coord(configs["back_to_search_pos"]["value"])
    message_input_pos = scale_coord(configs["message_input_pos"]["value"])

    try:
        orders_df = pd.read_excel(orders_file, dtype=str)
        orders = orders_df.iloc[:, 0].tolist()
    except FileNotFoundError:
        log(f"订单文件 {orders_file} 不存在", level='error')
        return
    except ValueError:
        log(f"订单文件格式错误", level='error')
        return
    except Exception as e:
        log(f"读取订单文件出错: {e}", level='error')
        return

    try:
        message_df = pd.read_excel(message_file)
        message_template = message_df.iloc[0, 0]
    except FileNotFoundError:
        log(f"消息文件 {message_file} 不存在", level='error')
        return
    except ValueError:
        log(f"消息文件格式错误", level='error')
        return
    except Exception as e:
        log(f"读取消息文件出错: {e}", level='error')
        return

    if not os.path.exists(log_file):
        wb = Workbook()
        ws = wb.active
        ws.append(['订单号', '状态', '描述', '时间'])
        wb.save(log_file)

    def log_step(order, status, description):
        try:
            wb = load_workbook(log_file)
            ws = wb.active
            timestamp = time.strftime('%Y-%m-%d %H:%M:%S')
            ws.append([order, status, description, timestamp])
            wb.save(log_file)
        except Exception as e:
            log(f"写入日志出错: {e}", level='error')

    def try_locate_image(image, retries=3, delay=1, confidence=0.7):
        for attempt in range(retries):
            try:
                loc = pyautogui.locateOnScreen(image, confidence=confidence)
                if loc:
                    return loc
                time.sleep(delay)
            except:
                log(f"图像未找到（尝试 {attempt + 1} 次），置信度={confidence}", level='error')
        return None

    i = 0
    if os.path.exists(progress_path):
        try:
            with open(progress_path, 'r') as f:
                i = json.load(f).get('index', 0)
        except Exception as e:
            log(f"加载进度文件失败: {e}", level='error')

    for idx, order in enumerate(orders[i:], start=i):
        pause_event.wait()
        if stop_event.is_set():
            with open(progress_path, 'w') as f:
                json.dump({'index': idx}, f)
            break
        current_status = '处理中'
        wait_time = 3 if idx == 0 else 1.5
        time.sleep(wait_time)
        try:
            pyautogui.click(*search_box_click_pos)
        except Exception as e:
            log_step(order, '失败', f'点击搜索框失败: {e}')
            continue

        if idx > 0:
            try:
                pyautogui.click(*search_box_click_pos)
                pyautogui.hotkey('ctrl', 'a')
                pyautogui.press('delete')
            except Exception as e:
                log_step(order, '失败', f'清除搜索框失败: {e}')
                continue

        try:
            clipboard.copy(str(order))
            pyautogui.hotkey('ctrl', 'v')
        except Exception as e:
            log_step(order, '失败', f'粘贴订单号失败: {e}')
            continue

        try:
            pyautogui.press('enter')
        except Exception as e:
            log_step(order, '失败', f'回车检索失败: {e}')
            continue

        if not wait_for_element(chat_button_img, timeout=4.5):
            log_step(order, '失败', '聊天页面加载超时')
            continue

        chat_button_loc = try_locate_image(chat_button_img, retries=3, delay=1, confidence=0.8)
        if chat_button_loc:
            try:
                pyautogui.click(pyautogui.center(chat_button_loc))
            except Exception as e:
                log_step(order, '失败', f'点击聊天按钮失败: {e}')
                continue
        else:
            log_step(order, '失败', '未识别到聊天按钮，跳过')
            continue

        time.sleep(8)

        no_send_loc1 = try_locate_image(no_send_img, retries=3, delay=1, confidence=0.7)
        no_send_loc2 = try_locate_image(no_send_img2, retries=3, delay=1, confidence=0.7)
        send_message_loc = try_locate_image(send_message_img, retries=3, delay=1, confidence=0.7)

        if send_message_loc:
            try:
                pyautogui.click(pyautogui.center(send_message_loc))
                clipboard.copy(message_template)
                pyautogui.hotkey('ctrl', 'v')
                pyautogui.press('enter')
                current_status = '发送成功'
                log_step(order, current_status, '消息发送成功')
            except Exception as e:
                log_step(order, '失败', f'发送消息失败: {e}')
        else:
            current_status = '无法发送'
            log_step(order, current_status, '未识别到消息输入框，跳过发送')

        if no_send_loc1 or no_send_loc2:
            log_step(order, current_status, f'检测到无法发送图片: fasongbuliao.png={bool(no_send_loc1)}, fasongbuliao2.png={bool(no_send_loc2)}')

        try:
            pyautogui.click(*back_to_search_pos)
        except Exception as e:
            log_step(order, '失败', f'返回搜索页面失败: {e}')

    if os.path.exists(progress_path):
        os.remove(progress_path)
    log('所有订单处理完成。日志已保存到运行日志.xlsx')

# ========== 运行模式函数 ==========
def start_mode(mode_func, mode_name):
    root.withdraw()

    status_win = tk.Toplevel()
    screen_height = root.winfo_screenheight()
    status_win.geometry('300x100+0+' + str(screen_height - 150))
    status_win.attributes('-topmost', True)

    tk.Label(status_win, text=f"正在运行 {mode_name} 模式").pack(pady=10)

    pause_event = Event()
    pause_event.set()
    stop_event = Event()
    thread_stop_event = Event()

    def toggle_pause():
        try:
            if not status_win.winfo_exists():
                return
            if pause_event.is_set():
                pause_event.clear()
                pause_btn.config(text='继续')
            else:
                pause_event.set()
                pause_btn.config(text='暂停')
        except tk.TclError as e:
            log(f"暂停按钮更新失败: {e}", level='error')

    def stop_mode():
        stop_event.set()
        thread_stop_event.set()
        status_win.destroy()
        root.deiconify()

    pause_btn = tk.Button(status_win, text='暂停', command=toggle_pause)
    pause_btn.pack(side='left', padx=20)
    tk.Button(status_win, text='结束', command=stop_mode).pack(side='right', padx=20)

    def handle_hotkeys():
        while not thread_stop_event.is_set():
            if keyboard.is_pressed('ctrl+c'):
                stop_mode()
                break
            elif keyboard.is_pressed('ctrl+p'):
                toggle_pause()
            time.sleep(0.1)

    thread = Thread(target=mode_func, args=(pause_event, stop_event))
    thread.start()
    Thread(target=handle_hotkeys).start()

    def check_thread():
        if not thread.is_alive():
            thread_stop_event.set()
            status_win.destroy()
            root.deiconify()
        else:
            status_win.after(1000, check_thread)

    check_thread()

# ========== Tkinter GUI：模式选择界面 ==========
root = tk.Tk()
root.title("自动化模式选择")
root.geometry("800x600")

# 实时鼠标位置
mouse_label = tk.Label(root, text="")
mouse_label.pack(anchor='w')

def update_mouse_label():
    try:
        x, y = pyautogui.position()
        mouse_label.config(text=f"当前鼠标位置: ({x}, {y})")
        root.after(100, update_mouse_label)
    except tk.TclError:
        pass

update_mouse_label()

notebook = ttk.Notebook(root)
notebook.pack(fill='both', expand=True)

modes = [
    ("查履约", run_mode_cha_lvyue),
    ("通过样品", run_mode_tongguo_yangpin),
    ("查到货", run_mode_cha_daohuo),
    ("催视频码", run_mode_cui_shipinma)
]

# 存储坐标标签以便更新
coord_labels = {}

for mode_name, mode_func in modes:
    frame = ttk.Frame(notebook)
    notebook.add(frame, text=mode_name)

    canvas = Canvas(frame, highlightthickness=0)
    scrollbar = Scrollbar(frame, orient="vertical", command=canvas.yview)
    scrollable_frame = ttk.Frame(canvas)

    scrollable_frame.bind(
        "<Configure>",
        lambda e, c=canvas: c.configure(scrollregion=c.bbox("all"))
    )

    canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
    canvas.configure(yscrollcommand=scrollbar.set)

    # 添加鼠标滚轮支持
    def on_mouse_wheel(event, c=canvas):
        c.yview_scroll(int(-1 * (event.delta / 120)), "units")
        return "break"  # 防止事件冒泡

    # 绑定滚轮事件到 canvas 和其子组件
    canvas.bind("<MouseWheel>", on_mouse_wheel)
    scrollable_frame.bind("<MouseWheel>", on_mouse_wheel)
    for child in scrollable_frame.winfo_children():
        child.bind("<MouseWheel>", on_mouse_wheel)

    # 确保 canvas 获得焦点以响应滚轮
    canvas.bind("<Enter>", lambda e: canvas.focus_set())
    canvas.bind("<Leave>", lambda e: root.focus_set())

    canvas.pack(side="left", fill="both", expand=True)
    scrollbar.pack(side="right", fill="y")

    # 在 scrollable_frame 内分成三列
    left_frame = ttk.Frame(scrollable_frame)
    middle_frame = ttk.Frame(scrollable_frame)
    right_frame = ttk.Frame(scrollable_frame)

    left_frame.pack(side="left", padx=10, pady=10, fill='both')
    middle_frame.pack(side="left", padx=10, pady=10, fill='both')
    right_frame.pack(side="left", padx=10, pady=10, fill='both')

    # 左列：坐标信息
    tk.Label(left_frame, text=f"{mode_name} 坐标信息：").pack(anchor='w')
    coord_labels[mode_name] = {}
    for coord_name, info in default_configs[mode_name].items():
        value = info["value"]
        comment = info["comment"]
        label = tk.Label(left_frame, text=f"{coord_name} = {value} {comment}")
        label.pack(anchor='w')
        coord_labels[mode_name][coord_name] = label
        tk.Button(left_frame, text="重设", command=lambda m=mode_name, c=coord_name, l=label: reset_coord(m, c, l)).pack(anchor='w')
        tk.Button(left_frame, text="测试", command=lambda m=mode_name, c=coord_name: pyautogui.click(*scale_coord(default_configs[m][c]["value"][:2]))).pack(anchor='w')
        # 绑定滚轮到每个子组件
        label.bind("<MouseWheel>", on_mouse_wheel)
        for btn in left_frame.winfo_children():
            if isinstance(btn, tk.Button):
                btn.bind("<MouseWheel>", on_mouse_wheel)

    # 中列：文件路径
    tk.Label(middle_frame, text=f"{mode_name} 文件路径：").pack(anchor='w')
    for file_info in mode_paths[mode_name]["files"]:
        tk.Label(middle_frame, text=f"{file_info['name']}: {file_info['path']}").pack(anchor='w')
        tk.Button(middle_frame, text="打开文件", command=lambda p=file_info['path']: os.startfile(p) if os.path.exists(p) else messagebox.showerror("错误", "文件不存在")).pack(anchor='w')
        # 绑定滚轮
        for widget in middle_frame.winfo_children():
            widget.bind("<MouseWheel>", on_mouse_wheel)

    # 右列：图片路径
    tk.Label(right_frame, text=f"{mode_name} 图片路径：").pack(anchor='w')
    def update_image_path(mode, img_name, entry):
        new_path = entry.get()
        for img in mode_paths[mode]["images"]:
            if img["name"] == img_name:
                img["path"] = new_path
                save_paths_and_configs()
                messagebox.showinfo("成功", f"已更新 {img_name} 路径为 {new_path}")
                break

    for img_info in mode_paths[mode_name]["images"]:
        tk.Label(right_frame, text=f"{img_info['name']}: ").pack(anchor='w')
        entry = Entry(right_frame)
        entry.insert(0, img_info['path'])
        entry.pack(anchor='w')
        tk.Button(right_frame, text="保存修改", command=lambda m=mode_name, n=img_info['name'], e=entry: update_image_path(m, n, e)).pack(anchor='w')
        # 绑定滚轮
        for widget in right_frame.winfo_children():
            widget.bind("<MouseWheel>", on_mouse_wheel)

    tk.Button(scrollable_frame, text=f"启动 {mode_name}", command=lambda f=mode_func, n=mode_name: start_mode(f, n)).pack(pady=20)
    # 绑定滚轮到启动按钮
    for widget in scrollable_frame.winfo_children():
        if isinstance(widget, tk.Button):
            widget.bind("<MouseWheel>", on_mouse_wheel)

# 验证路径并启动 GUI
validate_paths()
root.mainloop()