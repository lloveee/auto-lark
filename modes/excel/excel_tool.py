import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from typing import List, Tuple, Any, Union, Optional
from openpyxl.worksheet.worksheet import Worksheet
import os
from core.env import EXCEL_DIR
from pathlib import Path
import flet as ft
import uuid
import time


class ExcelTool:
    def __init__(
            self,
            default_headers: Optional[List[str]] = None,
            header_row: int = 1,
            file_name: Optional[str] = None
    ):
        if header_row < 1:
            raise ValueError("Header row must be 1 or greater.")
        self.default_headers = default_headers or ['序号', 'BD', '达人名称','带货店铺','样品名','合作类型','发样数量','性别','粉丝量','履约率','主页链接','佣金率','寄样批准日期','发货日期','订单号','履约方式','创作视频链接','视频码']
        self.header_row = header_row
        if not file_name:
            file_name = f"excel_{uuid.uuid4().hex[:8]}_{int(time.time())}.xlsx"
        self.file_path = EXCEL_DIR.joinpath(file_name)
        self.workbook = self._load_workbook(self.file_path) if os.path.exists(self.file_path) else self._create_workbook()

    def _create_workbook(self) -> openpyxl.Workbook:
        return openpyxl.Workbook()

    def _load_workbook(self, file_path: str) -> openpyxl.Workbook:
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"文件路径 {file_path} 不存在.")
        try:
            return openpyxl.load_workbook(file_path)
        except Exception as e:
            raise ValueError(f"加载文件失败: {str(e)}")

    def save(self) -> None:
        self.workbook.save(self.file_path)

    def get_sheet_names(self) -> List[str]:
        return self.workbook.sheetnames

    def get_sheet(self, sheet_name: str) -> openpyxl.worksheet.worksheet.Worksheet:
        if sheet_name not in self.workbook.sheetnames:
            raise KeyError(f"Sheet '{sheet_name}' 不存在")
        return self.workbook[sheet_name]

    def create_sheet(
            self,
            sheet_name: str,
            headers: Optional[List[str]] = None,
            position: Optional[int] = None,
            header_font: Optional[Font] = None,
            header_alignment: Optional[Alignment] = None,
            header_fill: Optional[PatternFill] = None
    ) -> openpyxl.worksheet.worksheet.Worksheet:
        sheet = self.workbook.create_sheet(sheet_name, position)
        headers_to_write = headers if headers is not None else self.default_headers
        if headers_to_write:
            for col, header in enumerate(headers_to_write, start=1):
                self.write_cell(
                    sheet, self.header_row, col, header,
                    font=header_font, alignment=header_alignment, fill=header_fill
                )
        return sheet

    def set_default_headers(self, headers: List[str]) -> None:
        self.default_headers = headers

    def check_headers(self, sheet: openpyxl.worksheet.worksheet.Worksheet,
                      expected_headers: Optional[List[str]] = None) -> bool:
        if self.header_row > sheet.max_row:
            raise ValueError(f"Header row {self.header_row} 超过了长度限制.")
        expected = expected_headers if expected_headers is not None else self.default_headers
        if not expected:
            return True
        actual_headers = [sheet.cell(row=self.header_row, column=col).value for col in range(1, len(expected) + 1)]
        return actual_headers == expected

    def write_cell(
            self,
            sheet: openpyxl.worksheet.worksheet.Worksheet,
            row: int,
            col: int,
            value: Any,
            font: Optional[Font] = None,
            alignment: Optional[Alignment] = None,
            fill: Optional[PatternFill] = None
    ) -> None:
        cell = sheet.cell(row=row, column=col)
        cell.value = value
        if font:
            cell.font = font
        if alignment:
            cell.alignment = alignment
        if fill:
            cell.fill = fill

    def read_cell(self, sheet: openpyxl.worksheet.worksheet.Worksheet, row: int, col: int) -> Any:
        return sheet.cell(row=row, column=col).value

    def append_row(self, sheet: openpyxl.worksheet.worksheet.Worksheet, data: List[Any]) -> None:
        sheet.append(data)

    def get_row_count(self, sheet: openpyxl.worksheet.worksheet.Worksheet) -> int:
        return sheet.max_row

    def get_column_count(self, sheet: openpyxl.worksheet.worksheet.Worksheet) -> int:
        return sheet.max_column

    def read_range(
            self,
            sheet: openpyxl.worksheet.worksheet.Worksheet,
            start_row: int,
            start_col: int,
            end_row: int,
            end_col: int,
            skip_headers: bool = True
    ) -> List[List[Any]]:
        if skip_headers and start_row <= self.header_row:
            start_row = self.header_row + 1
        return [
            [sheet.cell(row=row, column=col).value for col in range(start_col, end_col + 1)]
            for row in range(start_row, end_row + 1)
        ]

    def write_range(
            self,
            sheet: openpyxl.worksheet.worksheet.Worksheet,
            start_row: int,
            start_col: int,
            data: List[List[Any]]
    ) -> None:
        if start_row <= self.header_row:
            start_row = self.header_row + 1
        for row_idx, row_data in enumerate(data, start=start_row):
            for col_idx, value in enumerate(row_data, start=start_col):
                sheet.cell(row=row_idx, column=col_idx).value = value

    def set_column_width(self, sheet: openpyxl.worksheet.worksheet.Worksheet, column: int, width: float) -> None:
        column_letter = get_column_letter(column)
        sheet.column_dimensions[column_letter].width = width

    def set_row_height(self, sheet: openpyxl.worksheet.worksheet.Worksheet, row: int, height: float) -> None:
        sheet.row_dimensions[row].height = height

    def apply_format_to_range(
            self,
            sheet: openpyxl.worksheet.worksheet.Worksheet,
            start_row: int,
            start_col: int,
            end_row: int,
            end_col: int,
            font: Optional[Font] = None,
            alignment: Optional[Alignment] = None,
            fill: Optional[PatternFill] = None
    ) -> None:
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                cell = sheet.cell(row=row, column=col)
                if font:
                    cell.font = font
                if alignment:
                    cell.alignment = alignment
                if fill:
                    cell.fill = fill

    def preview_with_flet(self, page: ft.Page) -> ft.Control:
        sheet_dropdown = ft.Dropdown(
            label="选择工作表",
            options=[ft.dropdown.Option(sheet) for sheet in self.get_sheet_names()],
            value=self.get_sheet_names()[0] if self.get_sheet_names() else None
        )

        data_table = ft.DataTable(
            columns=[],
            rows=[],
            border=ft.border.all(1, ft.Colors.GREY_400),
            heading_row_color=ft.Colors.GREY_200,
            expand=True,
            column_spacing=10,
        )

        file_picker = ft.FilePicker(
            on_result=lambda e: on_file_picked(e)  # 文件选择后的回调
        )
        page.overlay.append(file_picker)

        def update_table(sheet_name: str):
            if not sheet_name:
                data_table.columns = []
                data_table.rows = []
                page.update()
                return

            self.workbook = self._load_workbook(self.file_path) if os.path.exists(
                self.file_path) else self._create_workbook()
            sheet = self.get_sheet(sheet_name)
            row_count = self.get_row_count(sheet)
            col_count = self.get_column_count(sheet)

            # Get headers (from default_headers or first row)
            headers = self.default_headers if self.check_headers(sheet) else [
                sheet.cell(row=self.header_row, column=col).value or f"Column {col}"
                for col in range(1, col_count + 1)
            ]

            # Update table columns
            data_table.columns = [
                ft.DataColumn(
                    ft.Text(
                        header,
                        weight=ft.FontWeight.BOLD,
                        text_align=ft.TextAlign.CENTER,  # 水平左对齐
                    )
                )
                for header in headers
            ]

            # Read data (excluding headers)
            if row_count > self.header_row:
                data = self.read_range(
                    sheet, self.header_row + 1, 1, row_count, col_count, skip_headers=True
                )
                data_table.rows = [
                    ft.DataRow(
                        cells=[
                            ft.DataCell(
                                ft.Text(
                                    str(cell) if cell is not None else "",
                                    text_align=ft.TextAlign.CENTER  # 数据行内容左对齐
                                ),
                            ) for cell in row
                        ]
                    )
                    for row in data
                ]
            else:
                data_table.rows = []

            page.update()

        def change_workbook(file_path: str):
            self.file_path = file_path
            self.workbook = self._load_workbook(self.file_path) if os.path.exists(
                self.file_path) else self._create_workbook()
            sheet_dropdown.options=[ft.dropdown.Option(sheet) for sheet in self.get_sheet_names()]
            sheet_dropdown.value=self.get_sheet_names()[0] if self.get_sheet_names() else None

            update_table(sheet_dropdown.value)

        def on_file_picked(e):
            if e.files and len(e.files) > 0:
                selected_file = e.files[0].path
                change_workbook(selected_file)
            else:
                page.snack_bar = ft.SnackBar(ft.Text("未选择文件"), open=True)
                page.update()

        # Initialize table with first sheet
        if sheet_dropdown.value:
            update_table(sheet_dropdown.value)

        # Refresh button
        refresh_button = ft.ElevatedButton(
            text="刷新",
            on_click=lambda e: update_table(sheet_dropdown.value)
        )

        select_file_button = ft.ElevatedButton(
            text="选择 需要查看的 Excel 文件",
            on_click=lambda e: file_picker.pick_files(
                allowed_extensions=["xlsx", "xls"],  # 限制为 Excel 文件
                allow_multiple=False
            ),
            icon=ft.Icons.FOLDER_OPEN
        )

        # 打开 Excel 按钮
        def open_excel_file(e):
            if os.path.exists(self.file_path):
                os.startfile(self.file_path)
            else:
                page.snack_bar = ft.SnackBar(ft.Text("Excel 文件不存在"), open=True)
                page.update()

        open_button = ft.ElevatedButton(
            text="打开 Excel 文件",
            on_click=open_excel_file,
            icon=ft.Icons.FILE_OPEN
        )


        scrollable_table = ft.Container(
            content=ft.Column(
                controls=[
                    ft.Row(
                        scroll=ft.ScrollMode.ALWAYS,  # 横向滚动
                        controls=[data_table],
                    )
                ],
                scroll=ft.ScrollMode.ALWAYS,  # 纵向滚动
                expand=True,
            )
        )


        # Return a Column containing the controls
        return ft.Column(
            controls=[
                ft.Row(
                    [sheet_dropdown, select_file_button, refresh_button, open_button],
                    alignment=ft.MainAxisAlignment.CENTER,
                    spacing=10
                ),
                ft.Container(
                    content=scrollable_table,
                    expand=True,
                    border=ft.border.all(1, ft.Colors.GREY_400),
                    border_radius=8,
                    padding=10,
                    alignment=ft.alignment.center,
                )
            ],
            expand=True,
            horizontal_alignment=ft.CrossAxisAlignment.STRETCH
        )


