import io

import flet as ft
from modes.excel.excel_tool import ExcelTool
import os
from core.logger import logger
from openpyxl import load_workbook

class YangPingPage(ft.Column):
    """信息录入"""

    def __init__(self, excel_tool: ExcelTool):
        super().__init__()
        self.excel_tool = excel_tool
        self.expand = True
        self.horizontal_alignment = ft.CrossAxisAlignment.CENTER
        self.alignment = ft.MainAxisAlignment.CENTER
        self.file_picker = None
        self.select_file_button = None
        self.sheet_dropdown = None

    def build(self):

        self.sheet_dropdown = ft.Dropdown(
            label="选择工作表",
            options=[ft.dropdown.Option(sheet) for sheet in self.excel_tool.get_sheet_names()],
            value=self.excel_tool.get_sheet_names()[0] if self.excel_tool.get_sheet_names() else None
        )

        self.file_picker = ft.FilePicker(on_result=self.on_file_picked)
        self.page.overlay.append(self.file_picker)
        self.select_file_button = ft.ElevatedButton(
            text="选择 需要查看的 Excel 文件",
            on_click=lambda e: self.file_picker.pick_files(
                allowed_extensions=["xlsx", "xls"],
                allow_multiple=False
            ),
            icon=ft.Icons.FOLDER_OPEN
        )

        config_section = ft.Container(
            content=ft.Column(
                controls=[
                    ft.Text("远程数据写入", size=20, weight=ft.FontWeight.BOLD),
                    ft.Row(
                        controls=[
                            self.sheet_dropdown,
                            self.select_file_button,
                        ],
                        spacing=20,
                        alignment=ft.MainAxisAlignment.CENTER,
                    ),

                ],
                spacing=10,
                horizontal_alignment=ft.CrossAxisAlignment.CENTER,
            ),
            padding=ft.padding.all(15),
            bgcolor=ft.Colors.WHITE,
            margin=ft.margin.only(bottom=15, top=10),
            shadow=ft.BoxShadow(
                spread_radius=0.5,
                blur_radius=4,
                color=ft.Colors.GREY_300,
                offset=ft.Offset(0, 2),
            ),
            border_radius=10,
        )

        self.controls = [
            config_section
        ]

    def on_file_picked(self, e):
        if not e.files:
            logger.warning("未选择文件")
        f = e.files[0]
        print(f"{f.name} => {f.path}")
        try:
            if f.path and os.path.exists(f.path):
                self.change_workbook_from_path(f.path)
            #web
            elif hasattr(f, "bytes") and f.bytes:
                self.change_workbook_from_bytes(f.name, f.bytes)
            else:
                logger.error("文件数据为空，打开失败")
        except Exception as e:
            logger.error(f"加载文件失败{e}")
            self.page.update()

    def change_workbook_from_path(self, file_path: str):
        try:
            old = self.excel_tool.file_path
            self.excel_tool.file_path = file_path
            logger.success(f"切换文件路径{old} => {file_path}")
            self.excel_tool.workbook = self.excel_tool.load_workbook(file_path) if os.path.exists(
                file_path) else self.excel_tool.create_workbook()
            self.refresh_dropdown()
        except Exception as e:
            logger.error(f"加载文件失败：{e}")

    def change_workbook_from_bytes(self, file_path: str, file_bytes: bytes):
        old = self.excel_tool.file_path
        self.excel_tool.file_path = file_path
        logger.success(f"切换文件路径 {old} => [内存文件: {file_path}]")
        wb = load_workbook(io.BytesIO(file_bytes))
        self.excel_tool.workbook = wb
        self.refresh_dropdown()

    def refresh_dropdown(self):
        sheet_names = self.excel_tool.get_sheet_names()
        self.sheet_dropdown.options = [ft.dropdown.Option(s) for s in sheet_names]
        self.sheet_dropdown.value = sheet_names[0] if sheet_names else None
        self.page.snack_bar = ft.SnackBar(ft.Text("文件加载成功！"), open=True)
        self.page.update()
